import time 
from openpyxl import Workbook 
wb= Workbook() 
# Starts new workbook (called wb)  

ws=wb.active 
# Pregrenerated worksheet in workbook 

ws1= wb.create_sheet("MySheet") 
#Format to create new sheet 
 
ws2 = wb.create_sheet("MySheet", 0)#Create worksheet at first position 
ws3 = wb.create_sheet("MySheet", 1)#Create worksheet at second position 
ws4 = wb.create_sheet("MySheet", -1)#Create worksheet at penultimate position
ws3.title= "Coronavirus"#Set new title of worksheet 
ws3 = wb["New Title"]#Once you gave a worksheet a name, you can get it as a key of the workbook
print(wb.sheetnames)#You can review the names of all worksheets of the workbook with the Workbook.sheetname attribute
target = wb.copy_worksheet(ws)#Copying worksheet

